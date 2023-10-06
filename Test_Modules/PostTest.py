from datetime import datetime  # to get today date
import shutil  # to move files
import openpyxl  # edit file excel
import PySimpleGUI as sg  # GUI window
from pathlib import Path  # make folder
import webbrowser  # open website
import pyautogui as pyautogui  # to screenshot monitor

def run(systemSN = 0, carrierSN = 0, leftBollard=0, rightBollard=0, operator=0, date=0, controlWalk=0, magnetwalk=0, objectWalk=0):
    
    #open excel file
    templatePath = 'Z:/05. Manufacturing/60. Uncontrolled/Troubleshoot/Dai/MIS Program/MIS-4251_Func_Test/727-4251 DataSheet.xlsx'
    wb = openpyxl.load_workbook(templatePath)
    ws = wb['Test Report']
    
    # Write Information
    ws['C3'] = systemSN
    ws['C4'] = leftBollard
    ws['C5'] = rightBollard
    ws['J3'] = operator
    ws['J5'] = date
    
    # Write the Test record
    ws['I10'] = "Yes"
    ws['I11'] = "Yes"
    ws['I12'] = "Yes"
    ws['I13'] = "Yes"
    ws['I14'] = "Yes"
    ws['I15'] = "Yes"
    ws['I16'] = "Yes"
    ws['I17'] = "Yes"
    ws['I18'] = "Yes"
    ws['I19'] = "Yes"
    ws['I20'] = "Yes"
    ws['I21'] = "Yes"
    ws['I22'] = "Yes"
    ws['I23'] = "Yes"
    ws['I24'] = "Yes"
    ws['I25'] = "Yes"
    ws['I26'] = "Yes"
    ws['I27'] = "Yes"
    ws['I28'] = "Yes"
    ws['I29'] = "Yes"
    ws['I30'] = "Yes"
    ws['I31'] = "Yes"
    ws['I32'] = "Yes"
    ws['I33'] = "Yes"
    ws['I34'] = "Yes"
    ws['I35'] = "Yes"
    ws['I36'] = "Yes"
    ws['I37'] = "Yes"
    ws['I38'] = "Yes"
    ws['I39'] = "Yes"
    ws['I40'] = "Yes"
    ws['I41'] = "Yes"
    ws['I42'] = "Yes"
    ws['I43'] = "Yes"
    ws['I44'] = "Yes"
    ws['I45'] = controlWalk
    ws['I46'] = magnetwalk
    ws['I47'] = "Yes"
    ws['I48'] = "Yes"
    ws['I49'] = objectWalk
    ws['I50'] = "Yes"
    ws['I51'] = "Yes"
    ws['I52'] = "Yes"
    ws['I53'] = "Yes"
    ws['I54'] = "Yes"
    ws['I55'] = "Yes"
    ws['I56'] = "Yes"
    
    #save excel file
    wb.save(str(systemSN) + " Datasheet.xlsx")
    src_folder = r"Z:\05. Manufacturing\60. Uncontrolled\Troubleshoot\Dai\MIS Program\MIS-4251_Func_Test\\"
    src1_folder = r"C:\Users\qtest1\Downloads\Download\\"
    src2_folder = r"Z:\05. Manufacturing\20. Test\400 records\Test Records\727\727-4251\\"
    dst_folder = r"Z:\05. Manufacturing\20. Test\400 records\Test Records\727\727-4251\Bollard " + str(systemSN) + "\\"
    file_name = str(systemSN) + " Datasheet.xlsx"
    
    shutil.move(src1_folder, dst_folder)
    shutil.move(src2_folder + carrierSN, dst_folder)
    shutil.move(src_folder + file_name, dst_folder)
    
    return True

if __name__ == "__main__":
    print("Debug Mode")
    run()